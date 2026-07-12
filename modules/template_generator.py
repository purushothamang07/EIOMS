from docxtpl import DocxTemplate
from openpyxl import load_workbook


def generate_word(template_path, output_path, data):

    doc = DocxTemplate(template_path)

    doc.render(data)

    doc.save(output_path)


def generate_excel(template_path, output_path, data):

    wb = load_workbook(template_path)

    for sheet in wb.worksheets:

        for row in sheet.iter_rows():

            for cell in row:

                if isinstance(cell.value, str):

                    for key, value in data.items():

                        placeholder = "{{" + key + "}}"

                        cell.value = cell.value.replace(
                            placeholder,
                            str(value)
                        )

    wb.save(output_path)


def generate_template(template_path, output_path, data):

    if template_path.lower().endswith(".docx"):

        generate_word(
            template_path,
            output_path,
            data
        )

    elif template_path.lower().endswith(".xlsx"):

        generate_excel(
            template_path,
            output_path,
            data
        )

    else:

        raise Exception("Unsupported template type")